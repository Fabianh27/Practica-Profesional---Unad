import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

# === 1. Cargar archivo original ===
# Cambia la ruta si el archivo no está en la misma carpeta del script
archivo = r"G:\Mi unidad\Universidad\PRACTICA PROFESIONAL\codigos power query\ANALISIS DETERIORO 2025.xlsx"

df = pd.read_excel(archivo)
df.columns = df.columns.str.strip()  # limpiar espacios en nombres de columnas

# === 2. Separar ingresos (1A) y salidas (2F, 2J, 2L) ===
ingresos = df[df["Comprobante"].str.upper() == "1A"]
salidas = df[df["Comprobante"].str.upper().isin(["2F", "2J", "2L"])]

# === 3. Detectar columnas que representan años ===
columnas_anos = [col for col in df.columns if col.isdigit()]

# === 4. Agrupar y sumar las salidas ===
salidas_sumadas = (
    salidas
    .groupby(["Codigo", "Descripcion de elementos"], as_index=False)
    .agg({**{col: "sum" for col in columnas_anos}, "Comprobante": "count"})
)

# === 5. Calcular la suma total de los años (Total_Anios) ===
salidas_sumadas["Total"] = salidas_sumadas[columnas_anos].sum(axis=1)

# === 6. Renombrar y agregar columnas ===
salidas_sumadas = salidas_sumadas.rename(columns={"Comprobante": "Conteo_Salidas"})
salidas_sumadas["Comprobante"] = "2F_2J_2L"
salidas_sumadas["Movimiento"] = "SALIDAS"

# === 7. Agregar conteo y total a los ingresos ===
ingresos["Conteo_Salidas"] = 0
ingresos["Total"] = ingresos[columnas_anos].sum(axis=1)

# === 8. Reorganizar las columnas ===
columnas_finales = (
    ["Codigo", "Comprobante", "Movimiento", "Descripcion de elementos", "Conteo_Salidas"]
    + columnas_anos
    + ["Total"]
)

final_df = pd.concat([ingresos[columnas_finales], salidas_sumadas[columnas_finales]], ignore_index=True)

# === 9. Exportar a Excel ===
output_file = "ANALISIS_DETERIORO_2025_modificado.xlsx"
final_df.to_excel(output_file, index=False)

# === 10. Aplicar formato visual con openpyxl ===
wb = load_workbook(output_file)
ws = wb.active

# --- Estilo del encabezado ---
header_fill = PatternFill(start_color="79ccb3", end_color="79ccb3", fill_type="solid")
header_font = Font(bold=True, color="000000")
border = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
align_center = Alignment(horizontal="center", vertical="center")

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = align_center
    cell.border = border

# --- Ajustar ancho de columnas automáticamente ---
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# --- Agregar tabla con filtros y franjas ---
ref = f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"
table = Table(displayName="Tabla_Deterioro", ref=ref)
style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False)
table.tableStyleInfo = style
ws.add_table(table)

# --- Guardar archivo final ---
wb.save(output_file)

print("✅ Archivo creado con formato:", output_file)
print("📊 Incluye columnas 'Conteo_Salidas' y 'Total' (suma total de los años 2001–2024).")
